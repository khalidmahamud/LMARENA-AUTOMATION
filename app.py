"""LM Arena Side-by-Side Automation — Entry Point.

Run:  python app.py
Open: http://localhost:8000
"""

from __future__ import annotations

import asyncio
import base64
import csv
import io
import json as json_mod
import logging
import re
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from functools import partial
from pathlib import Path

from fastapi import FastAPI, File, UploadFile, WebSocket
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

from src.browser.manager import BrowserManager
from src.browser.selectors import SelectorRegistry
from src.checkpoint.manager import CheckpointManager
from src.core.events import EventBus
from src.export.excel_exporter import export_to_csv, export_to_excel, export_to_json
from src.models.config import AppConfig
from src.models.results import RunResult, WindowResult
from src.orchestrator.run_orchestrator import RunOrchestrator
from src.proxy.pool import ProxyPool
from src.proxy.xlsx_source import load_proxy_candidates_from_xlsx
from src.transport.ws_broadcaster import WsBroadcaster
from src.transport.ws_handler import WsHandler
from src.preview.screenshot_service import ScreenshotService

# ── Logging ──

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Globals (wired at startup) ──

config: AppConfig
event_bus: EventBus
broadcaster: WsBroadcaster
browser_manager: BrowserManager
checkpoint_manager: CheckpointManager
ws_handler: WsHandler
screenshot_service: ScreenshotService
proxy_pool: ProxyPool


def _load_proxy_source(protocol: str = "http", limit: int | None = None) -> list[dict]:
    """Load proxy candidates from the configured XLSX source."""
    return load_proxy_candidates_from_xlsx(
        config.proxy_source_xlsx,
        protocol=protocol,
        limit=limit,
    )


@asynccontextmanager
async def lifespan(application: FastAPI):
    """Startup / shutdown lifecycle."""
    global config, event_bus, broadcaster, browser_manager, checkpoint_manager, ws_handler, proxy_pool, screenshot_service

    config = AppConfig.from_yaml("config/default_config.yaml")
    SelectorRegistry.load("config/selectors.yaml")

    # Initialize proxy pool from the XLSX source only.
    proxy_pool = ProxyPool(check_fn=_check_proxy, source_loader=_load_proxy_source)
    source_exists = Path(config.proxy_source_xlsx).exists()
    if not source_exists:
        logger.warning("Proxy source XLSX not found: %s", config.proxy_source_xlsx)
    elif source_exists:
        # Populate and check the pool directly from the XLSX source on startup.
        asyncio.create_task(proxy_pool.health_check_all())
        logger.info(
            "Background health check started from source=%s",
            config.proxy_source_xlsx,
        )

    event_bus = EventBus()
    broadcaster = WsBroadcaster(event_bus)
    browser_manager = BrowserManager(config, proxy_pool=proxy_pool)
    checkpoint_manager = CheckpointManager(config.output_dir)
    await browser_manager.start()
    screenshot_service = ScreenshotService(browser_manager, config.preview)
    await screenshot_service.start()

    def orchestrator_factory() -> RunOrchestrator:
        return RunOrchestrator(
            config, event_bus, browser_manager, checkpoint_manager
        )

    ws_handler = WsHandler(orchestrator_factory, broadcaster, checkpoint_manager, screenshot_service=screenshot_service)

    logger.info("Server ready — open http://localhost:8000")
    yield

    await proxy_pool.stop_auto_refresh()
    await screenshot_service.stop()
    await browser_manager.close_all()
    logger.info("Shutdown complete")


app = FastAPI(title="LM Arena Automation", lifespan=lifespan)
app.mount("/static", StaticFiles(directory="static"), name="static")


# ── Routes ──


@app.get("/", response_class=HTMLResponse)
async def index():
    return Path("templates/index.html").read_text()


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await ws_handler.handle(websocket)


@app.post("/upload-prompts")
async def upload_prompts(file: UploadFile):
    """Parse a CSV or Excel file and return columns + all rows."""
    if not file.filename:
        return JSONResponse({"error": "No file provided"}, status_code=400)

    ext = Path(file.filename).suffix.lower()
    raw = await file.read()

    try:
        if ext == ".csv":
            text = raw.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            columns = reader.fieldnames or []
            rows = [dict(r) for r in reader]
        elif ext == ".xlsx":
            wb = load_workbook(filename=io.BytesIO(raw), read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return JSONResponse({"error": "Empty spreadsheet"}, status_code=400)
            columns = [str(c) if c else f"Column {i+1}" for i, c in enumerate(all_rows[0])]
            rows = [
                {columns[j]: (str(cell) if cell is not None else "")
                 for j, cell in enumerate(row)}
                for row in all_rows[1:]
            ]
            wb.close()
        else:
            return JSONResponse(
                {"error": f"Unsupported file type: {ext}. Use .csv or .xlsx"},
                status_code=400,
            )
    except Exception as exc:
        logger.error("Failed to parse upload: %s", exc)
        return JSONResponse({"error": f"Parse error: {exc}"}, status_code=400)

    return {
        "filename": file.filename,
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
        "preview": rows[:5],
    }


INSTRUCTION_FIELDS = {
    "prompt", "turns", "images", "window_count",
    "submission_gap_seconds", "model_a", "model_b", "retain_output",
    "clear_cookies", "incognito", "simultaneous_start", "zoom_pct",
}

# Pattern for numbered turn columns: prompt_1, prompt_2, images_1, etc.
_TURN_COLUMN_RE = re.compile(r"^(prompt|images)_(\d+)$")

BOOL_FIELDS = {"clear_cookies", "incognito", "simultaneous_start"}
INT_FIELDS = {"window_count": (1, 12), "zoom_pct": (25, 200)}
FLOAT_FIELDS = {"submission_gap_seconds": (5.0, 300.0)}

ALLOWED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
_MIME_MAP = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
    ".gif": "image/gif",
}


def _build_image_pool(
    image_files: list[tuple[str, bytes]],
) -> dict[str, dict]:
    """Build ``{filename: {data, mime_type, filename}}`` from uploaded images."""
    pool: dict[str, dict] = {}
    for filename, data in image_files:
        ext = Path(filename).suffix.lower()
        mime = _MIME_MAP.get(ext)
        if not mime:
            continue
        pool[filename] = {
            "data": base64.b64encode(data).decode("ascii"),
            "mime_type": mime,
            "filename": filename,
        }
    return pool


def _resolve_image_refs(
    filenames: list[str],
    image_pool: dict[str, dict],
) -> list[dict] | None:
    """Resolve image filename references against the pool (case-insensitive)."""
    resolved = []
    for fname in filenames:
        entry = image_pool.get(fname)
        if not entry:
            fname_lower = fname.lower()
            for pool_name, pool_entry in image_pool.items():
                if pool_name.lower() == fname_lower:
                    entry = pool_entry
                    break
        if entry:
            resolved.append(entry)
    return resolved if resolved else None


def _coerce_instruction(
    raw: dict,
    image_pool: dict[str, dict] | None = None,
) -> dict:
    """Validate and coerce a raw instruction dict."""
    inst: dict = {}
    turn_columns: dict[int, dict] = {}  # {N: {"prompt": ..., "images": ...}}

    for k, v in raw.items():
        key = k.strip().lower().replace(" ", "_")

        # Check for numbered turn columns (prompt_1, images_2, etc.)
        turn_match = _TURN_COLUMN_RE.match(key)
        if turn_match:
            col_type = turn_match.group(1)  # "prompt" or "images"
            col_num = int(turn_match.group(2))
            if col_num > 10:
                continue  # cap at 10 turns
            if v is not None and (not isinstance(v, str) or v.strip()):
                turn_columns.setdefault(col_num, {})[col_type] = v
            continue

        if key not in INSTRUCTION_FIELDS:
            continue
        if v is None or (isinstance(v, str) and not v.strip()):
            continue

        # Pass-through for JSON "turns" array
        if key == "turns":
            if isinstance(v, list):
                inst["turns"] = v
            continue

        # Pass-through for "images" (list of filenames or dicts)
        if key == "images":
            if isinstance(v, list):
                inst["images"] = v
            elif isinstance(v, str) and v.strip():
                # CSV: comma-separated filenames
                inst["images"] = [f.strip() for f in v.split(",") if f.strip()]
            continue

        if key in BOOL_FIELDS:
            if isinstance(v, str):
                v = v.strip().lower() in ("true", "1", "yes")
            inst[key] = bool(v)
        elif key in INT_FIELDS:
            lo, hi = INT_FIELDS[key]
            inst[key] = max(lo, min(hi, int(float(v))))
        elif key in FLOAT_FIELDS:
            lo, hi = FLOAT_FIELDS[key]
            inst[key] = max(lo, min(hi, float(v)))
        else:
            inst[key] = str(v).strip() if isinstance(v, str) else v

    # Assemble turns from numbered columns (CSV/Excel: prompt_1, prompt_2, …)
    if turn_columns and "turns" not in inst:
        turns = []
        for n in sorted(turn_columns.keys()):
            col = turn_columns[n]
            text = str(col.get("prompt", "")).strip()
            if not text:
                continue
            turn_entry: dict = {"text": text}
            img_refs = col.get("images")
            if img_refs and image_pool:
                fnames = [f.strip() for f in str(img_refs).split(",") if f.strip()]
                resolved = _resolve_image_refs(fnames, image_pool)
                if resolved:
                    turn_entry["images"] = resolved
            turns.append(turn_entry)
        if turns:
            inst["turns"] = turns

    # Resolve image filename refs inside JSON turns
    if "turns" in inst and image_pool:
        for turn in inst["turns"]:
            if isinstance(turn, dict) and "images" in turn:
                imgs = turn["images"]
                if isinstance(imgs, list) and imgs and isinstance(imgs[0], str):
                    resolved = _resolve_image_refs(imgs, image_pool)
                    turn["images"] = resolved

    # Resolve image filename refs for single-turn "images" field
    if "images" in inst and "turns" not in inst and image_pool:
        imgs = inst["images"]
        if isinstance(imgs, list) and imgs and isinstance(imgs[0], str):
            resolved = _resolve_image_refs(imgs, image_pool)
            inst["images"] = resolved if resolved else None

    return inst


@app.post("/upload-instructions")
async def upload_instructions(files: list[UploadFile] = File(...)):
    """Parse instruction file(s) with optional image files.

    Accepts one instruction file (``.json`` / ``.csv`` / ``.xlsx``) and any
    number of image files (``.png`` / ``.jpg`` / ``.webp`` / ``.gif``).
    Image filenames referenced in the instruction file are resolved against
    the co-uploaded images.
    """
    if not files:
        return JSONResponse({"error": "No files provided"}, status_code=400)

    # Separate instruction file from image files
    instruction_file: UploadFile | None = None
    image_files: list[tuple[str, bytes]] = []

    for f in files:
        if not f.filename:
            continue
        ext = Path(f.filename).suffix.lower()
        if ext in {".json", ".csv", ".xlsx"}:
            if instruction_file is None:
                instruction_file = f
        elif ext in ALLOWED_IMAGE_EXTENSIONS:
            raw = await f.read()
            image_files.append((f.filename, raw))

    if instruction_file is None:
        return JSONResponse(
            {"error": "No instruction file found (.json, .csv, or .xlsx)"},
            status_code=400,
        )

    # Build image pool from co-uploaded images
    image_pool = _build_image_pool(image_files) if image_files else None

    ext = Path(instruction_file.filename).suffix.lower()
    raw_bytes = await instruction_file.read()

    try:
        if ext == ".json":
            data = json_mod.loads(raw_bytes.decode("utf-8-sig"))
            if not isinstance(data, list):
                return JSONResponse(
                    {"error": "JSON must be an array of instruction objects"},
                    status_code=400,
                )
            raw_rows = data
        elif ext == ".csv":
            text = raw_bytes.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            raw_rows = [dict(r) for r in reader]
        elif ext == ".xlsx":
            wb = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return JSONResponse({"error": "Empty spreadsheet"}, status_code=400)
            columns = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(all_rows[0])]
            raw_rows = [
                {columns[j]: cell for j, cell in enumerate(row)}
                for row in all_rows[1:]
            ]
            wb.close()
        else:
            return JSONResponse(
                {"error": f"Unsupported file type: {ext}. Use .json, .csv, or .xlsx"},
                status_code=400,
            )
    except Exception as exc:
        logger.error("Failed to parse instruction file: %s", exc)
        return JSONResponse({"error": f"Parse error: {exc}"}, status_code=400)

    # Validate and coerce
    instructions = []
    for i, raw in enumerate(raw_rows):
        inst = _coerce_instruction(raw, image_pool=image_pool)
        if not inst.get("prompt") and not inst.get("turns"):
            continue  # skip rows without prompt or turns
        instructions.append(inst)

    if not instructions:
        return JSONResponse(
            {"error": "No valid instructions found (each row needs a 'prompt' or 'turns')"},
            status_code=400,
        )

    return {
        "filename": instruction_file.filename,
        "instructions": instructions,
        "count": len(instructions),
    }


@app.get("/export")
async def export_excel(run_id: str | None = None, scope: str | None = None):
    result = _resolve_export_result(run_id=run_id, scope=scope)
    if result:
        path = export_to_excel(result, config.output_dir)
        return FileResponse(
            path,
            filename=path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return {"error": "No results available"}


@app.get("/export-csv")
async def export_csv_file(run_id: str | None = None, scope: str | None = None):
    result = _resolve_export_result(run_id=run_id, scope=scope)
    if result:
        path = export_to_csv(result, config.output_dir)
        return FileResponse(path, filename=path.name, media_type="text/csv")
    return {"error": "No results available"}


@app.get("/export-json")
async def export_json(run_id: str | None = None, scope: str | None = None):
    result = _resolve_export_result(run_id=run_id, scope=scope)
    if result:
        path = export_to_json(result, config.output_dir)
        return FileResponse(
            path,
            filename=path.name,
            media_type="application/json",
        )
    return {"error": "No results available"}


def _resolve_export_result(
    run_id: str | None = None,
    scope: str | None = None,
) -> RunResult | None:
    """Resolve export scope.

    - run_id set: export that run only
    - scope=all: export merged results of every run that has results
    - default: export the latest run
    """
    if run_id:
        orch = ws_handler.get_orchestrator(run_id)
        return orch.last_result if orch and orch.last_result else None

    if (scope or "").strip().lower() == "all":
        results = [
            orch.last_result
            for orch in ws_handler.get_all_orchestrators().values()
            if orch and orch.last_result
        ]
        if not results:
            return None
        results.sort(key=lambda r: r.started_at)
        if len(results) == 1:
            return results[0]
        return _merge_run_results(results)

    orch = ws_handler.orchestrator
    return orch.last_result if orch and orch.last_result else None


def _merge_run_results(results: list[RunResult]) -> RunResult:
    """Merge multiple runs into one exportable payload."""
    merged_rows: list[WindowResult] = []
    merged_prompts: list[str] = []
    for run in results:
        merged_prompts.extend(run.prompts or [run.prompt])
        for row in run.window_results:
            merged_rows.append(
                row.model_copy(
                    update={"run_id": row.run_id or run.run_id}
                )
            )

    successful = sum(1 for r in merged_rows if r.success)
    failed = len(merged_rows) - successful
    started_at = min(r.started_at for r in results)
    completed_candidates = [r.completed_at for r in results if r.completed_at]
    completed_at = max(completed_candidates) if completed_candidates else None
    total_elapsed = sum(r.total_elapsed_seconds or 0.0 for r in results)

    return RunResult(
        run_id=f"combined_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}",
        prompt=(
            results[0].prompt
            if len(results) == 1
            else f"Combined export ({len(results)} runs)"
        ),
        prompts=merged_prompts,
        total_batches=sum(r.total_batches for r in results),
        started_at=started_at,
        completed_at=completed_at,
        total_elapsed_seconds=total_elapsed,
        window_results=merged_rows,
        total_windows=len(merged_rows),
        successful_windows=successful,
        failed_windows=failed,
    )


# ── Run state endpoint ──


@app.get("/api/run-state")
async def get_run_state():
    """Return current run state for UI sync on reconnect."""
    state = ws_handler.get_run_state()
    if state:
        return state
    return {"running": False}


@app.post("/api/toggle-headless")
async def toggle_headless(enabled: bool = True):
    """Toggle headless mode at runtime."""
    config.browser.headless = enabled
    return {"headless": config.browser.headless}


@app.post("/api/preview/open-window")
async def open_preview_window(
    worker_id: int,
    run_id: str | None = None,
):
    """Bring a previewed browser window to the front when available."""
    return await browser_manager.focus_window(
        worker_id=worker_id,
        run_id=run_id,
    )


@app.post("/api/close-all-windows")
async def close_all_windows():
    """Stop active runs and close every currently open browser window."""
    await ws_handler._handle_stop_run()
    await browser_manager.close_open_windows()
    return {"ok": True}


TEST_URL = "https://arena.ai/"


def _check_proxy(proxy_server: str, timeout: int = 5) -> float:
    """Test if a proxy can HTTPS-tunnel to arena.ai. Returns latency in ms, or -1.0 on failure."""
    import socket
    import ssl
    import time as _time

    start = _time.monotonic()
    try:
        # Parse proxy host:port
        stripped = proxy_server.split("://")[-1]
        proxy_host, proxy_port = stripped.rsplit(":", 1)
        proxy_port = int(proxy_port)

        if proxy_server.startswith("socks"):
            # SOCKS: try PySocks, fall back to TCP connect
            try:
                import socks as pysocks
                proto = pysocks.SOCKS5 if "socks5" in proxy_server else pysocks.SOCKS4
                s = pysocks.socksocket(socket.AF_INET, socket.SOCK_STREAM)
                s.set_proxy(proto, proxy_host, proxy_port)
                s.settimeout(timeout)
                s.connect(("arena.ai", 443))
                # Try TLS handshake
                ctx = ssl.create_default_context()
                ss = ctx.wrap_socket(s, server_hostname="arena.ai")
                ss.close()
                return (_time.monotonic() - start) * 1000.0
            except ImportError:
                sock = socket.create_connection((proxy_host, proxy_port), timeout=timeout)
                sock.close()
                return (_time.monotonic() - start) * 1000.0
        else:
            # HTTP proxy: send CONNECT request (this is exactly what Chromium does)
            sock = socket.create_connection((proxy_host, proxy_port), timeout=timeout)
            connect_req = f"CONNECT arena.ai:443 HTTP/1.1\r\nHost: arena.ai:443\r\n\r\n"
            sock.sendall(connect_req.encode())

            response = b""
            while b"\r\n\r\n" not in response:
                chunk = sock.recv(4096)
                if not chunk:
                    sock.close()
                    return -1.0
                response += chunk

            # Check for 200 Connection established
            status_line = response.split(b"\r\n")[0].decode(errors="ignore")
            if "200" not in status_line:
                sock.close()
                return -1.0

            # TLS handshake through the tunnel — proves it actually works
            ctx = ssl.create_default_context()
            ss = ctx.wrap_socket(sock, server_hostname="arena.ai")
            ss.close()
            return (_time.monotonic() - start) * 1000.0
    except Exception:
        return -1.0


@app.get("/api/free-proxies")
async def fetch_free_proxies(
    protocol: str = "http", limit: int = 10, test: bool = False
):
    """Load proxies from the configured XLSX source. If test=true, health-check each proxy."""
    allowed = {"http", "https", "socks4", "socks5"}
    if protocol not in allowed:
        return JSONResponse(
            {"error": f"Invalid protocol. Choose from: {', '.join(sorted(allowed))}"},
            status_code=400,
        )
    limit = max(1, min(limit, 100))

    try:
        source_limit = limit if not test else max(limit * 10, 50)
        data = await asyncio.get_event_loop().run_in_executor(
            None,
            partial(_load_proxy_source, protocol, source_limit),
        )
    except Exception as exc:
        logger.error("Failed to load proxies from XLSX source: %s", exc)
        return JSONResponse({"error": f"Failed to load proxies: {exc}"}, status_code=502)

    if not test:
        proxies = data[:limit]
        return {
            "proxies": proxies,
            "count": len(proxies),
            "protocol": protocol,
            "tested": False,
            "source": config.proxy_source_xlsx,
        }

    candidates = data[:source_limit]
    candidate_servers = [entry["server"] for entry in candidates if entry.get("server")]
    logger.info("Testing %d proxies to find %d working ones...", len(candidates), limit)

    # Check in batches of 20 to avoid overwhelming the network
    loop = asyncio.get_event_loop()
    results = []
    batch_size = 20
    for i in range(0, len(candidate_servers), batch_size):
        batch = candidate_servers[i : i + batch_size]
        batch_results = await asyncio.gather(
            *(loop.run_in_executor(None, _check_proxy, proxy) for proxy in batch)
        )
        results.extend(batch_results)

    alive = [
        {
            "server": entry["server"],
            "username": entry.get("username"),
            "password": entry.get("password"),
            "latency_ms": round(latency, 1),
        }
        for entry, latency in zip(candidates, results)
        if latency > 0
    ]
    alive.sort(key=lambda p: p["latency_ms"])
    alive = alive[:limit]
    logger.info("Found %d working proxies out of %d tested", len(alive), len(candidates))

    # Also add to proxy pool
    if alive:
        proxy_pool.add_proxies(alive, source="xlsx")

    return {
        "proxies": alive,
        "count": len(alive),
        "protocol": protocol,
        "tested": True,
        "total_tested": len(candidates),
        "source": config.proxy_source_xlsx,
    }


# ── Proxy pool endpoints ──


@app.get("/api/proxy-pool/status")
async def proxy_pool_status():
    """Return current pool status for UI display."""
    try:
        return proxy_pool.get_status()
    except Exception as exc:
        logger.error("proxy_pool_status error: %s", exc)
        return {
            "total": 0,
            "healthy": 0,
            "unhealthy": 0,
            "auto_refresh_enabled": False,
            "auto_refresh_active": False,
            "proxies": [],
        }


@app.post("/api/proxy-pool/add")
async def add_to_proxy_pool(request: dict):
    """Manually add proxies to the pool. Body: {"proxies": [...]}"""
    proxies = request.get("proxies", [])
    added = proxy_pool.add_proxies(proxies, source="manual")
    return {"added": added, "total": proxy_pool.total_count}


@app.post("/api/proxy-pool/auto-refresh/start")
async def start_auto_refresh(
    protocol: str = "http", limit: int = 20, interval: int = 300
):
    """Start the background auto-refresh task."""
    await proxy_pool.start_auto_refresh(
        protocol=protocol, fetch_limit=limit, interval=float(interval)
    )
    return {"started": True, "interval": interval}


@app.post("/api/proxy-pool/auto-refresh/stop")
async def stop_auto_refresh():
    """Stop the background auto-refresh task."""
    await proxy_pool.stop_auto_refresh()
    return {"stopped": True}


@app.post("/api/proxy-pool/max-size")
async def set_pool_max_size(limit: int = 50):
    """Update the max healthy proxies cap."""
    proxy_pool.set_max_healthy(limit)
    return {"max_healthy": limit}


@app.post("/api/proxy-pool/max-latency")
async def set_pool_max_latency(ms: int = 5000):
    """Update the max latency threshold in ms."""
    proxy_pool.set_max_latency(float(ms))
    return {"max_latency_ms": ms}


@app.post("/api/proxy-pool/health-check")
async def health_check_pool():
    """Trigger an immediate health check of all proxies in the pool."""
    stats = await proxy_pool.health_check_all()
    return stats


# ── Checkpoint endpoints ──


@app.get("/api/checkpoints")
async def list_checkpoints():
    """Return all resumable (in_progress) checkpoint summaries."""
    checkpoints = checkpoint_manager.list_resumable()
    return [
        {
            "run_id": cp.run_id,
            "total_prompts": len(cp.all_prompts),
            "completed_prompts": len(cp.completed_prompt_indices),
            "next_batch": cp.next_batch_index,
            "total_batches": cp.total_batches,
            "last_checkpoint_at": cp.last_checkpoint_at,
            "status": cp.status,
        }
        for cp in checkpoints
    ]


@app.delete("/api/checkpoints/{run_id}")
async def delete_checkpoint(run_id: str):
    """Discard a checkpoint."""
    checkpoint_manager.delete(run_id)
    return {"deleted": run_id}


# ── Entry point ──

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000, ws_max_size=67_108_864)
