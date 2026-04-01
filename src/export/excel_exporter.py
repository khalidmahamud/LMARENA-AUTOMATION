from __future__ import annotations

import csv
import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models.results import RunResult

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


_CARD_RUN_ID_PATTERN = re.compile(r"^card_[^_]+_(\d+)$")


def _prompt_number_from_run_id(run_id: str | None) -> str:
    if not run_id:
        return ""
    match = _CARD_RUN_ID_PATTERN.match(run_id)
    if match:
        return match.group(1)
    return ""


def _run_label_from_run_id(run_id: str | None) -> str:
    prompt_no = _prompt_number_from_run_id(run_id)
    if prompt_no:
        return f"P#{prompt_no}"
    if not run_id:
        return ""
    return run_id if len(run_id) <= 16 else f"{run_id[:13]}..."


def export_to_excel(run_result: RunResult, output_dir: str = "outputs") -> Path:
    """Generate an ``.xlsx`` file from a ``RunResult``.

    Returns the ``Path`` to the written file.
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    filename = out / f"arena_results_{run_result.run_id}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    has_batches = run_result.total_batches > 1
    has_run_ids = any(wr.run_id for wr in run_result.window_results)
    has_prompt_numbers = any(
        _prompt_number_from_run_id(wr.run_id)
        for wr in run_result.window_results
    )
    has_turns = any(wr.turn_index > 0 for wr in run_result.window_results)

    # Header row
    headers = ["Window #"]
    if has_prompt_numbers:
        headers.append("Prompt #")
    if has_run_ids:
        headers.append("Run")
        headers.append("Run ID")
    if has_batches:
        headers.append("Batch")
    if has_turns:
        headers.append("Turn")
    headers += [
        "Prompt",
        "Model A",
        "Response A",
        "Model B",
        "Response B",
        "Elapsed (s)",
        "Status",
        "Error",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Data rows
    for row_idx, wr in enumerate(run_result.window_results, start=2):
        c = 1
        ws.cell(row=row_idx, column=c, value=wr.worker_id + 1)
        c += 1
        if has_prompt_numbers:
            ws.cell(
                row=row_idx,
                column=c,
                value=_prompt_number_from_run_id(wr.run_id),
            )
            c += 1
        if has_run_ids:
            ws.cell(
                row=row_idx,
                column=c,
                value=_run_label_from_run_id(wr.run_id),
            )
            c += 1
            ws.cell(row=row_idx, column=c, value=wr.run_id or "")
            c += 1
        if has_batches:
            ws.cell(row=row_idx, column=c, value=wr.batch_index + 1)
            c += 1
        if has_turns:
            ws.cell(row=row_idx, column=c, value=wr.turn_index + 1)
            c += 1
        ws.cell(row=row_idx, column=c, value=wr.prompt or "").alignment = WRAP_ALIGNMENT
        c += 1
        ws.cell(row=row_idx, column=c, value=wr.model_a_name or "")
        c += 1
        ws.cell(row=row_idx, column=c, value=wr.model_a_response or "").alignment = WRAP_ALIGNMENT
        c += 1
        ws.cell(row=row_idx, column=c, value=wr.model_b_name or "")
        c += 1
        ws.cell(row=row_idx, column=c, value=wr.model_b_response or "").alignment = WRAP_ALIGNMENT
        c += 1
        ws.cell(row=row_idx, column=c, value=round(wr.elapsed_seconds, 1) if wr.elapsed_seconds else "")
        c += 1
        ws.cell(row=row_idx, column=c, value="success" if wr.success else "error")
        c += 1
        ws.cell(row=row_idx, column=c, value=wr.error or "")

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Run ID"
    summary["B1"] = run_result.run_id

    # Show all prompts if they differ, otherwise just the single prompt
    unique_prompts = list(dict.fromkeys(run_result.prompts)) if run_result.prompts else [run_result.prompt]
    if len(unique_prompts) > 1:
        summary["A2"] = "Prompts"
        summary["B2"] = "\n".join(
            f"#{i+1}: {p[:200]}" for i, p in enumerate(unique_prompts)
        )
        summary["B2"].alignment = WRAP_ALIGNMENT
    else:
        summary["A2"] = "Prompt"
        summary["B2"] = run_result.prompt[:1000]

    summary["A3"] = "Total Batches"
    summary["B3"] = run_result.total_batches
    summary["A4"] = "Total Prompts"
    summary["B4"] = run_result.total_windows
    summary["A5"] = "Successful"
    summary["B5"] = run_result.successful_windows
    summary["A6"] = "Failed"
    summary["B6"] = run_result.failed_windows
    summary["A7"] = "Total Time (s)"
    summary["B7"] = (
        round(run_result.total_elapsed_seconds, 1)
        if run_result.total_elapsed_seconds
        else ""
    )

    # Column widths
    width_map = {
        "Window #": 10,
        "Prompt #": 10,
        "Run": 9,
        "Run ID": 24,
        "Batch": 8,
        "Turn": 8,
        "Prompt": 40,
        "Model A": 25,
        "Response A": 50,
        "Model B": 25,
        "Response B": 50,
        "Elapsed (s)": 12,
        "Status": 10,
        "Error": 30,
    }
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width_map.get(
            header, 20
        )

    wb.save(str(filename))
    logger.info("Excel exported to %s", filename)
    return filename


def export_to_csv(run_result: RunResult, output_dir: str = "outputs") -> Path:
    """Generate a ``.csv`` file from a ``RunResult``."""
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    filename = out / f"arena_results_{run_result.run_id}.csv"

    has_batches = run_result.total_batches > 1
    has_run_ids = any(wr.run_id for wr in run_result.window_results)
    has_prompt_numbers = any(
        _prompt_number_from_run_id(wr.run_id)
        for wr in run_result.window_results
    )
    has_turns = any(wr.turn_index > 0 for wr in run_result.window_results)
    headers = ["window"]
    if has_prompt_numbers:
        headers.append("prompt_no")
    if has_run_ids:
        headers.append("run")
        headers.append("run_id")
    if has_batches:
        headers.append("batch")
    if has_turns:
        headers.append("turn")
    headers += [
        "prompt", "model_a", "response_a",
        "model_b", "response_b", "elapsed_s", "status", "error",
    ]

    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for wr in run_result.window_results:
            row = [wr.worker_id + 1]
            if has_prompt_numbers:
                row.append(_prompt_number_from_run_id(wr.run_id))
            if has_run_ids:
                row.append(_run_label_from_run_id(wr.run_id))
                row.append(wr.run_id or "")
            if has_batches:
                row.append(wr.batch_index + 1)
            if has_turns:
                row.append(wr.turn_index + 1)
            row += [
                wr.prompt or "",
                wr.model_a_name or "",
                wr.model_a_response or "",
                wr.model_b_name or "",
                wr.model_b_response or "",
                round(wr.elapsed_seconds, 1) if wr.elapsed_seconds else "",
                "success" if wr.success else "error",
                wr.error or "",
            ]
            writer.writerow(row)

    logger.info("CSV exported to %s", filename)
    return filename


def export_to_json(run_result: RunResult, output_dir: str = "outputs") -> Path:
    """Generate a ``.json`` file from a ``RunResult``."""
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    filename = out / f"arena_results_{run_result.run_id}.json"

    data = run_result.model_dump(mode="json")
    for wr in data.get("window_results", []):
        run_id = wr.get("run_id")
        wr["prompt_no"] = _prompt_number_from_run_id(run_id)
        wr["run"] = _run_label_from_run_id(run_id)
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    logger.info("JSON exported to %s", filename)
    return filename
