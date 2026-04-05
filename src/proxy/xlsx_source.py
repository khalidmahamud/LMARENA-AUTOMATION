"""Load proxy candidates from a local XLSX source file."""

from __future__ import annotations

import os
import re
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

IP_PORT_RE = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}:\d{2,5}$")

_SERVER_HEADERS = {
    "proxy",
    "proxy_server",
    "proxy_url",
    "server",
    "url",
}
_HOST_HEADERS = {"host", "ip", "ip_address", "address"}
_PORT_HEADERS = {"port"}
_PROTOCOL_HEADERS = {"protocol", "scheme", "type"}
_USERNAME_HEADERS = {"username", "user"}
_PASSWORD_HEADERS = {"password", "pass"}
_LATENCY_HEADERS = {"latency_ms", "latency", "ping", "ping_ms"}
_CHECKED_AT_HEADERS = {"checked_at", "last_checked", "timestamp"}


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return text.replace(" ", "_").replace("-", "_")


def _normalize_protocol(value: object) -> str:
    text = str(value or "").strip().lower()
    if text == "https":
        return "https"
    if text == "socks4":
        return "socks4"
    if text == "socks5":
        return "socks5"
    return "http"


def _infer_protocol(server: str, fallback: Optional[str] = None) -> str:
    if "://" in server:
        return _normalize_protocol(server.split("://", 1)[0])
    return _normalize_protocol(fallback)


def _build_server(
    raw_server: object,
    host: object,
    port: object,
    row_protocol: object,
    fallback_protocol: Optional[str],
) -> Optional[str]:
    server = str(raw_server or "").strip()
    if server:
        if "://" in server:
            return server
        if IP_PORT_RE.match(server):
            return f"{_normalize_protocol(row_protocol or fallback_protocol)}://{server}"
        return None

    host_text = str(host or "").strip()
    port_text = str(port or "").strip()
    if not host_text or not port_text:
        return None
    if not port_text.isdigit():
        return None
    protocol = _normalize_protocol(row_protocol or fallback_protocol)
    return f"{protocol}://{host_text}:{port_text}"


def load_proxy_candidates_from_xlsx(
    path: str | Path,
    protocol: Optional[str] = None,
    limit: Optional[int] = None,
    sort_by_latency: bool = False,
) -> List[dict]:
    """Return de-duplicated proxy dicts from the configured XLSX file."""
    xlsx_path = Path(path)
    if not xlsx_path.exists():
        return []

    requested_protocol = _normalize_protocol(protocol)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []

        header_index: Dict[str, int] = {
            _normalize_header(value): idx for idx, value in enumerate(header_row)
        }

        server_idx = next(
            (header_index[name] for name in _SERVER_HEADERS if name in header_index),
            None,
        )
        host_idx = next(
            (header_index[name] for name in _HOST_HEADERS if name in header_index),
            None,
        )
        port_idx = next(
            (header_index[name] for name in _PORT_HEADERS if name in header_index),
            None,
        )
        protocol_idx = next(
            (header_index[name] for name in _PROTOCOL_HEADERS if name in header_index),
            None,
        )
        username_idx = next(
            (header_index[name] for name in _USERNAME_HEADERS if name in header_index),
            None,
        )
        password_idx = next(
            (header_index[name] for name in _PASSWORD_HEADERS if name in header_index),
            None,
        )
        latency_idx = next(
            (header_index[name] for name in _LATENCY_HEADERS if name in header_index),
            None,
        )

        if server_idx is None and (host_idx is None or port_idx is None):
            return []

        proxies: List[dict] = []
        seen = set()

        for row in rows:
            raw_protocol = row[protocol_idx] if protocol_idx is not None and protocol_idx < len(row) else None
            server = _build_server(
                row[server_idx] if server_idx is not None and server_idx < len(row) else None,
                row[host_idx] if host_idx is not None and host_idx < len(row) else None,
                row[port_idx] if port_idx is not None and port_idx < len(row) else None,
                raw_protocol,
                requested_protocol,
            )
            if not server or server in seen:
                continue

            entry_protocol = _infer_protocol(server, raw_protocol or requested_protocol)
            if protocol and entry_protocol != requested_protocol:
                continue

            proxy = {"server": server}
            if username_idx is not None and username_idx < len(row) and row[username_idx]:
                proxy["username"] = str(row[username_idx]).strip()
            if password_idx is not None and password_idx < len(row) and row[password_idx]:
                proxy["password"] = str(row[password_idx]).strip()

            # Read latency from XLSX so it can be used for sorting and pre-loading
            if latency_idx is not None and latency_idx < len(row):
                raw_lat = row[latency_idx]
                try:
                    proxy["latency_ms"] = float(raw_lat) if raw_lat is not None else None
                except (ValueError, TypeError):
                    pass

            seen.add(server)
            proxies.append(proxy)

            # When sorting, we need all rows first; otherwise break early
            if not sort_by_latency and limit is not None and len(proxies) >= max(1, int(limit)):
                break

        if sort_by_latency:
            proxies.sort(key=lambda p: p.get("latency_ms") or float("inf"))
        if limit is not None:
            proxies = proxies[: max(1, int(limit))]
        return proxies
    finally:
        wb.close()


def write_back_latencies_to_xlsx(
    path: str | Path,
    updates: Dict[str, Dict],
) -> int:
    """Write updated latency_ms and checked_at values back to the XLSX file.

    Args:
        path: Path to the xlsx file.
        updates: Mapping of proxy server URL to {"latency_ms": float, "checked_at": str}.

    Returns:
        Number of rows updated.
    """
    xlsx_path = Path(path)
    if not xlsx_path.exists() or not updates:
        return 0

    wb = load_workbook(xlsx_path)
    try:
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        header_index = {_normalize_header(v): idx for idx, v in enumerate(header_row)}

        server_idx = next(
            (header_index[n] for n in _SERVER_HEADERS if n in header_index), None
        )
        latency_idx = next(
            (header_index[n] for n in _LATENCY_HEADERS if n in header_index), None
        )
        checked_at_idx = next(
            (header_index[n] for n in _CHECKED_AT_HEADERS if n in header_index), None
        )

        if server_idx is None or latency_idx is None:
            return 0

        # Build a lookup stripping protocol prefixes for flexible matching
        stripped_lookup: Dict[str, Dict] = {}
        for key, upd in updates.items():
            stripped = key.split("://", 1)[-1] if "://" in key else key
            stripped_lookup[stripped] = upd
            stripped_lookup[key] = upd

        updated = 0
        for row_num in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_num, column=server_idx + 1).value
            if not cell_val:
                continue
            server_key = str(cell_val).strip()
            matched = stripped_lookup.get(server_key)
            if matched is None:
                # Also try stripping protocol from the cell value
                stripped_cell = server_key.split("://", 1)[-1] if "://" in server_key else server_key
                matched = stripped_lookup.get(stripped_cell)
            if matched is None:
                continue

            if "latency_ms" in matched and matched["latency_ms"] is not None:
                ws.cell(row=row_num, column=latency_idx + 1).value = matched["latency_ms"]
            if checked_at_idx is not None and "checked_at" in matched:
                ws.cell(row=row_num, column=checked_at_idx + 1).value = matched["checked_at"]
            updated += 1

        # Atomic write: save to temp file then rename
        dir_path = xlsx_path.parent
        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=dir_path)
        os.close(fd)
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, xlsx_path)
        except Exception:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
            raise

        return updated
    finally:
        wb.close()
