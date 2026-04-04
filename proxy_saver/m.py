"""
Continuous Proxy Harvester
==========================
Fetches free proxies from Proxifly CDN, checks latency in batches,
and appends good ones to an Excel file. Never deletes existing rows.

Usage:  python proxy_harvester.py
Configure everything in the CONFIG section at the bottom of this file.
"""

import json
import os
import socket
import time
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

import ssl

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("openpyxl not installed. Run:  pip install openpyxl")
    raise

try:
    import socks as pysocks
    HAS_PYSOCKS = True
except ImportError:
    HAS_PYSOCKS = False
    print("[!] PySocks not installed. SOCKS proxy checking will use basic TCP test. Run:  pip install pysocks")

# ─────────────────────────────────────────────
#  Core functions
# ─────────────────────────────────────────────

import re

IP_PORT_RE = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}:\d{2,5}$")


def normalize_source_protocol(value: str | None) -> str:
    """Normalize protocol labels from source URLs / payloads."""
    text = str(value or "").strip().lower()
    if "socks5" in text:
        return "socks5"
    if "socks4" in text:
        return "socks4"
    return "http"


def infer_source_protocol(url: str) -> str:
    """Best-effort protocol detection from a source URL."""
    return normalize_source_protocol(url)


def load_source_urls(txt_path: str) -> list[str]:
    """Read proxy source URLs from a text file. Skips comments and blanks."""
    path = Path(txt_path)
    if not path.exists():
        print(f"[!] proxy-list.txt not found at {path}")
        return []
    urls = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        urls.append(line)
    print(f"[+] Loaded {len(urls)} source URLs from {path.name}")
    return urls


def _parse_response(raw: bytes, url: str) -> list[dict]:
    """Auto-detect JSON vs TXT and return unified list of proxy metadata."""
    text = raw.decode("utf-8", errors="ignore").strip()
    proxies = []
    source_protocol = infer_source_protocol(url)

    # Try JSON first
    if text.startswith("[") or text.startswith("{"):
        try:
            data = json.loads(text)
            if isinstance(data, dict):
                # Some APIs wrap in {"data": [...]} or {"proxies": [...]}
                for key in ("data", "proxies", "results"):
                    if key in data and isinstance(data[key], list):
                        data = data[key]
                        break
                else:
                    data = [data]
            if isinstance(data, list):
                for entry in data:
                    if isinstance(entry, dict):
                        # proxifly format: {"proxy": "...", "geolocation": {...}}
                        proxy = entry.get("proxy") or entry.get("ip") or entry.get("host")
                        port = entry.get("port")
                        if proxy and port and ":" not in str(proxy):
                            proxy = f"{proxy}:{port}"
                        if proxy:
                            geo = entry.get("geolocation", {})
                            country = ""
                            if isinstance(geo, dict):
                                country = geo.get("country", "")
                            elif isinstance(entry.get("country"), str):
                                country = entry["country"]
                            entry_protocol = normalize_source_protocol(
                                entry.get("protocol")
                                or entry.get("type")
                                or entry.get("scheme")
                                or entry.get("protocols")
                                or source_protocol
                            )
                            proxies.append({
                                "proxy": str(proxy).strip(),
                                "country": country,
                                "protocol": entry_protocol,
                            })
                    elif isinstance(entry, str) and IP_PORT_RE.match(entry.strip()):
                        proxies.append({
                            "proxy": entry.strip(),
                            "country": "",
                            "protocol": source_protocol,
                        })
            if proxies:
                return proxies
        except (json.JSONDecodeError, ValueError):
            pass

    # Fallback: plain text — one ip:port per line
    for line in text.splitlines():
        line = line.strip()
        if IP_PORT_RE.match(line):
            proxies.append({
                "proxy": line,
                "country": "",
                "protocol": source_protocol,
            })

    return proxies


def fetch_all_sources(source_urls: list[str], allowed_protocols: set[str]) -> list[dict]:
    """Fetch proxies from all source URLs. Returns unified list."""
    all_proxies = []
    seen = set()
    for url in source_urls:
        req = urllib.request.Request(url, headers={"User-Agent": "ProxyHarvester/1.0"})
        try:
            with urllib.request.urlopen(req, timeout=20) as resp:
                raw = resp.read()
            entries = _parse_response(raw, url)
            new = 0
            skipped_protocol = 0
            for entry in entries:
                if entry.get("protocol", "http") not in allowed_protocols:
                    skipped_protocol += 1
                    continue
                p = entry["proxy"]
                if p not in seen:
                    seen.add(p)
                    all_proxies.append(entry)
                    new += 1
            print(
                f"  [+] {url[:70]}...  ->  {len(entries)} found, "
                f"{new} new, {skipped_protocol} skipped (protocol mismatch)"
            )
        except Exception as e:
            print(f"  [!] {url[:70]}...  ->  FAILED: {e}")
    print(f"[+] Total unique proxies from all sources: {len(all_proxies)}")
    return all_proxies


def _check_socks_latency(proxy_url: str, timeout: int, protocol: str) -> float:
    """Check SOCKS proxy latency via socket connection + TLS handshake."""
    addr = proxy_url.split("://")[-1]
    host, _, port_str = addr.rpartition(":")
    if not host or not port_str:
        return -1.0
    try:
        port = int(port_str)
    except ValueError:
        return -1.0

    start = time.perf_counter()
    try:
        if HAS_PYSOCKS:
            proto = pysocks.SOCKS5 if protocol == "socks5" else pysocks.SOCKS4
            s = pysocks.socksocket(socket.AF_INET, socket.SOCK_STREAM)
            s.set_proxy(proto, host, port)
            s.settimeout(timeout)
            s.connect(("arena.ai", 443))
            ctx = ssl.create_default_context()
            ss = ctx.wrap_socket(s, server_hostname="arena.ai")
            ss.close()
        else:
            # Fallback: just test TCP connectivity to the proxy itself
            sock = socket.create_connection((host, port), timeout=timeout)
            sock.close()
        return round((time.perf_counter() - start) * 1000, 1)
    except Exception:
        return -1.0


def check_proxy_latency(proxy_url: str, test_url: str, timeout: int, protocol: str = "http") -> float:
    """Check a single proxy's latency. Returns latency in ms, or -1 on failure."""
    if protocol in ("socks4", "socks5"):
        return _check_socks_latency(proxy_url, timeout, protocol)
    # HTTP / HTTPS — use urllib ProxyHandler
    proxy_handler = urllib.request.ProxyHandler({
        "http": proxy_url,
        "https": proxy_url,
    })
    opener = urllib.request.build_opener(proxy_handler)
    opener.addheaders = [("User-Agent", "ProxyHarvester/1.0")]
    start = time.perf_counter()
    try:
        with opener.open(test_url, timeout=timeout) as resp:
            resp.read()
        elapsed_ms = (time.perf_counter() - start) * 1000
        return round(elapsed_ms, 1)
    except Exception:
        return -1.0


def check_batch(
    proxies: list[tuple[str, str]],
    test_url: str,
    timeout: int,
    max_workers: int,
) -> list[tuple[str, str, float]]:
    """Check a batch of (proxy_url, protocol) concurrently. Returns list of (proxy, protocol, latency_ms)."""
    results = []
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        future_map = {
            pool.submit(check_proxy_latency, p, test_url, timeout, proto): (p, proto)
            for p, proto in proxies
        }
        for future in as_completed(future_map):
            proxy, proto = future_map[future]
            try:
                latency = future.result()
            except Exception:
                latency = -1.0
            results.append((proxy, proto, latency))
    return results


def load_existing_proxies(excel_path: str) -> set[str]:
    """Read existing proxy servers from the Excel file to avoid duplicates."""
    path = Path(excel_path)
    if not path.exists():
        return set()
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        existing = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                existing.add(str(row[0]).strip())
        wb.close()
        return existing
    except Exception as e:
        print(f"[!] Could not read existing Excel file: {e}")
        return set()


def append_to_excel(excel_path: str, rows: list[dict]) -> int:
    """Append new proxy rows to the Excel file. Creates file if missing. Returns count added."""
    if not rows:
        return 0

    path = Path(excel_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Proxies"
        headers = ["proxy_server", "latency_ms", "protocol", "country", "checked_at"]
        ws.append(headers)
        # Bold headers
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

    for row in rows:
        ws.append([
            row["proxy_server"],
            row["latency_ms"],
            row["protocol"],
            row.get("country", ""),
            row["checked_at"],
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    wb.save(path)
    wb.close()
    return len(rows)


def run_harvest_cycle(config: dict) -> dict:
    """Run one full harvest cycle: fetch -> check -> save. Returns stats."""
    protocols_cfg = config["protocols"]
    if isinstance(protocols_cfg, str):
        protocols_cfg = [protocols_cfg]
    allowed_protocols = {p.strip().lower() for p in protocols_cfg}
    batch_size = config["batch_size"]
    max_latency_ms = config["max_latency_ms"]
    timeout_sec = config["proxy_timeout_sec"]
    max_workers = config["max_workers"]
    test_url = config["test_url"]
    excel_path = config["excel_path"]
    source_file = config["proxy_list_file"]

    stats = {"fetched": 0, "checked": 0, "passed": 0, "new_added": 0, "duplicates_skipped": 0}

    print(f"[*] Checking protocols: {', '.join(sorted(allowed_protocols))}")
    if test_url.lower().startswith("https://") and "http" in allowed_protocols:
        print("[!] HTTPS test URLs are stricter and usually yield fewer passing free HTTP proxies.")

    # 1. Load source URLs from proxy-list.txt
    source_urls = load_source_urls(source_file)
    if not source_urls:
        print("[!] No source URLs found. Add URLs to proxy-list.txt")
        return stats

    # 2. Fetch proxies from all sources
    raw_list = fetch_all_sources(source_urls, allowed_protocols)
    if not raw_list:
        return stats

    # 3. Build candidate list with protocol info
    candidates = []          # list of (proxy_url, protocol)
    candidate_meta = {}      # proxy_url -> {country, protocol}
    for entry in raw_list:
        proxy = entry.get("proxy", "")
        entry_protocol = entry.get("protocol", "http")
        if proxy and proxy not in candidate_meta:
            # Add protocol prefix if it's raw ip:port
            if IP_PORT_RE.match(proxy):
                proxy_url = f"{entry_protocol}://{proxy}"
            else:
                proxy_url = proxy
            candidates.append((proxy_url, entry_protocol))
            candidate_meta[proxy_url] = {"country": entry.get("country", ""), "protocol": entry_protocol}
    stats["fetched"] = len(candidates)
    print(f"[*] {len(candidates)} unique proxies to check")

    # 4. Load existing to skip duplicates
    existing = load_existing_proxies(excel_path)
    print(f"[*] {len(existing)} proxies already in Excel")

    # 5. Check in batches — save to Excel after EVERY batch
    for batch_start in range(0, len(candidates), batch_size):
        batch = candidates[batch_start : batch_start + batch_size]
        batch_num = (batch_start // batch_size) + 1
        total_batches = (len(candidates) + batch_size - 1) // batch_size
        print(f"\n--- Batch {batch_num}/{total_batches} ({len(batch)} proxies) ---")

        results = check_batch(batch, test_url, timeout_sec, max_workers)
        stats["checked"] += len(results)

        batch_rows = []
        for proxy, proto, latency in results:
            if latency <= 0 or latency > max_latency_ms:
                continue
            stats["passed"] += 1
            if proxy in existing:
                stats["duplicates_skipped"] += 1
                continue
            existing.add(proxy)
            batch_rows.append({
                "proxy_server": proxy,
                "latency_ms": latency,
                "protocol": proto,
                "country": candidate_meta.get(proxy, {}).get("country", ""),
                "checked_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            })

        # Save this batch's good proxies immediately
        added = append_to_excel(excel_path, batch_rows)
        stats["new_added"] += added

        good_in_batch = sum(1 for _, _, lat in results if 0 < lat <= max_latency_ms)
        print(f"    Passed: {good_in_batch}/{len(batch)} | Saved: {added} new IPs")

    return stats


def run_continuous(config: dict):
    """Main loop — runs harvest cycles forever with a sleep between them."""
    cycle = 0
    protocols = config["protocols"] if isinstance(config["protocols"], list) else [config["protocols"]]
    print("=" * 60)
    print("  PROXY HARVESTER — Continuous Mode")
    print(f"  Protocols: {', '.join(protocols)}")
    print(f"  Sources: {config['proxy_list_file']}")
    print(f"  Excel: {config['excel_path']}")
    print(f"  Max Latency: {config['max_latency_ms']}ms")
    print(f"  Batch Size: {config['batch_size']}  |  Interval: {config['cycle_interval_sec']}s")
    print("=" * 60)

    while True:
        cycle += 1
        print(f"\n{'='*60}")
        print(f"  CYCLE {cycle}  —  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}")

        try:
            stats = run_harvest_cycle(config)
            print(f"\n[RESULT] Fetched: {stats['fetched']} | Checked: {stats['checked']} | "
                  f"Passed (<{config['max_latency_ms']}ms): {stats['passed']} | "
                  f"New Added: {stats['new_added']} | Duplicates Skipped: {stats['duplicates_skipped']}")
        except Exception as e:
            print(f"\n[ERROR] Cycle {cycle} failed: {e}")

        wait = config["cycle_interval_sec"]
        print(f"\n[*] Sleeping {wait}s until next cycle...")
        try:
            time.sleep(wait)
        except KeyboardInterrupt:
            print("\n[!] Stopped by user. Goodbye!")
            break


# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION — Edit these values to customize the harvester
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":

    CONFIG = {
        # Path to proxy-list.txt (add/remove source URLs there)
        "proxy_list_file": os.path.join(os.path.dirname(__file__), "proxy-list.txt"),

        # Protocols to check: any combination of "http", "socks4", "socks5"
        "protocols": ["http", "socks4", "socks5"],

        # Max acceptable latency in milliseconds
        "max_latency_ms": 6500,

        # How many proxies to check in one concurrent batch
        "batch_size": 1000,

        # Max concurrent threads per batch (careful with too high values)
        "max_workers": 1000,

        # Timeout per proxy check in seconds
        "proxy_timeout_sec": 15,

        # URL used to test proxy connectivity & measure latency
        "test_url": "https://arena.ai/",

        # Seconds to wait between full harvest cycles
        "cycle_interval_sec": 300,

        # Path to the output Excel file (relative or absolute)
        "excel_path": os.path.join(os.path.dirname(__file__), "good_proxies.xlsx"),
    }

    run_continuous(CONFIG)
